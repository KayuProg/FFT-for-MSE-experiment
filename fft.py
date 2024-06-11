import numpy as np
import matplotlib.pyplot as plt
import openpyxl as op

num=1
c="m"
sheetname= f"3-{num}{c}"


wb=op.load_workbook("FFT.xlsx")
wb_s=op.load_workbook("FFT_s.xlsx")

ws=wb[sheetname]
ws_s=wb_s[sheetname]

n=8096#データ数
f=1/1000#サンプリング周波数

values=[]
values_s=[]

for i in range(n):
    val=ws[f"A{i+1}"].value
    val=float(val)
    values.append(val)

    val_s=ws_s[f"A{i+1}"].value
    val_s=float(val_s)
    values_s.append(val_s)


fft=np.fft.fft(values)
freq=np.fft.fftfreq(n,f)
fft=fft/(n/2)#正規化
amp=np.abs(fft)

fft_s=np.fft.fft(values_s)
fft_s=fft_s/(n/2)
amp_s=np.abs(fft_s)



plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'
plt.rcParams['xtick.major.width'] = 2.0
plt.rcParams['ytick.major.width'] = 2.0
plt.rcParams['font.size'] = 20
plt.rcParams['axes.linewidth'] = 1.0
plt.rcParams["legend.framealpha"] = 1
plt.rcParams["figure.figsize"] = [10,7]
plt.rcParams['font.family'] = 'serif'
plt.rcParams['font.serif'] = ['Times New Roman'] + plt.rcParams['font.serif']
plt.xlim(0,100)
# # plt.ylim(top=0.003)
plt.xlabel("Frequency [Hz]",fontsize='20')
plt.ylabel("Amplitude",fontsize='20')
plt.tick_params(labelsize=15)
plt.grid(linestyle='solid', alpha=0.5)
plt.plot(freq[1:int(n/2)], amp[1:int(n/2)],color='black')
plt.plot(freq[1:int(n/2)], amp_s[1:int(n/2)],color='red')

if c=="t":
    plt.plot(freq[1:int(n / 2)], amp[1:int(n / 2)], color='black',
             label=f"Pole s{num} angle FFT data")
    plt.plot(freq[1:int(n / 2)], amp_s[1:int(n / 2)], color='red',
             label=f"Pole s{num} simulation angle FFT data")
elif c=="m":
    plt.plot(freq[1:int(n / 2)], amp[1:int(n / 2)], color='black',
             label=f"Pole s{num} displacement FFT data")
    plt.plot(freq[1:int(n / 2)], amp_s[1:int(n / 2)], color='red',
             label=f"Pole s{num} simulation displacement FFT data")

plt.legend(loc="upper right")

plt.savefig(f"./fig/{sheetname}.png")
plt.show()
